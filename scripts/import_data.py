#!/usr/bin/env python3
"""
Import Chittety Construction Excel data into JSON files and seed the SQLite database.
"""
import json
import sqlite3
import os

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl

UPLOAD_DIR = "/home/z/my-project/upload"
DATA_DIR = "/home/z/my-project/src/lib/data"
os.makedirs(DATA_DIR, exist_ok=True)

MASTER_FILE = f"{UPLOAD_DIR}/chittety_master_catalog_1500_products_services_database.xlsx"

def import_products_and_services():
    wb = openpyxl.load_workbook(MASTER_FILE, read_only=True, data_only=True)
    
    # === PRODUCTS ===
    ws = wb['Products']
    products = []
    categories_map = {}  # code_prefix -> category info
    subcategories_set = set()
    brands_set = set()
    
    # Read Category Architecture
    ws_cat = wb['Category Architecture']
    for row in ws_cat.iter_rows(min_row=2, values_only=True):
        if row[0]:
            categories_map[row[0]] = {
                'code': row[0],
                'name': row[1],
                'skuRange': row[2],
                'menuLabel': row[3],
                'cta': row[4]
            }
    
    # Read products (skip header rows 1-3)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:  # Skip empty SKU rows
            continue
        
        sku = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ''
        category = str(row[2]).strip() if row[2] else ''
        subcategory = str(row[3]).strip() if row[3] else ''
        brand = str(row[4]).strip() if row[4] else ''
        product_type = str(row[5]).strip() if row[5] else ''
        spec = str(row[6]).strip() if row[6] else ''
        unit = str(row[7]).strip() if row[7] else 'Each'
        moq = int(row[8]) if row[8] else 1
        market_price = float(row[9]) if row[9] else 0
        discount = float(row[10]) if row[10] else 0.1
        # Chittety price = market_price * (1 - discount)
        chittety_price = round(market_price * (1 - discount), 2)
        availability = str(row[12]).strip() if row[12] else 'Vendor Available / Confirm Stock'
        price_basis = str(row[13]).strip() if row[13] else 'Market reference'
        short_desc = str(row[22]).strip() if row[22] else ''
        
        # Extract images
        images = []
        for i in range(14, 21):  # columns 15-21 (0-indexed 14-20)
            if row[i] and str(row[i]).strip():
                images.append(str(row[i]).strip())
        
        # Determine code prefix from SKU
        code_prefix = sku.split('-')[1] if '-' in sku else 'UNK'
        
        products.append({
            'sku': sku,
            'name': name,
            'category': category,
            'subcategory': subcategory,
            'brand': brand,
            'productType': product_type,
            'specification': spec,
            'unit': unit,
            'moq': moq,
            'marketPrice': market_price,
            'discount': discount,
            'chittetyPrice': chittety_price,
            'availability': availability,
            'priceBasis': price_basis,
            'shortDescription': short_desc,
            'images': images,
            'codePrefix': code_prefix
        })
        
        if subcategory:
            subcategories_set.add(f"{category}|||{subcategory}")
        if brand:
            brands_set.add(brand)
    
    # Build subcategories map
    subcategories_map = {}
    for item in sorted(subcategories_set):
        cat, sub = item.split('|||')
        if cat not in subcategories_map:
            subcategories_map[cat] = []
        subcategories_map[cat].append(sub)
    
    # === SERVICES ===
    ws_svc = wb['Services Pricing']
    services = []
    service_categories = set()
    
    for row in ws_svc.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        
        code = str(row[0]).strip()
        category = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        starting_price = float(row[3]) if row[3] else 0
        price_type = str(row[4]).strip() if row[4] else 'Starting From'
        notes = str(row[5]).strip() if row[5] else ''
        cta = str(row[6]).strip() if row[6] else ''
        source = str(row[7]).strip() if row[7] else ''
        
        services.append({
            'code': code,
            'category': category,
            'name': name,
            'startingPrice': starting_price,
            'priceType': price_type,
            'notes': notes,
            'cta': cta,
            'source': source
        })
        service_categories.add(category)
    
    wb.close()
    
    # Save JSON files
    with open(f"{DATA_DIR}/products.json", 'w') as f:
        json.dump(products, f)
    
    with open(f"{DATA_DIR}/categories.json", 'w') as f:
        json.dump(list(categories_map.values()), f)
    
    with open(f"{DATA_DIR}/subcategories.json", 'w') as f:
        json.dump(subcategories_map, f)
    
    with open(f"{DATA_DIR}/brands.json", 'w') as f:
        json.dump(sorted(list(brands_set)), f)
    
    with open(f"{DATA_DIR}/services.json", 'w') as f:
        json.dump(services, f)
    
    with open(f"{DATA_DIR}/service-categories.json", 'w') as f:
        json.dump(sorted(list(service_categories)), f)
    
    print(f"Imported {len(products)} products")
    print(f"Imported {len(categories_map)} categories")
    print(f"Imported {len(subcategories_set)} subcategories")
    print(f"Imported {len(brands_set)} brands")
    print(f"Imported {len(services)} services")
    print(f"Service categories: {sorted(service_categories)}")
    
    # === SEED SQLITE DATABASE ===
    db_path = "/home/z/my-project/db/chittety.db"
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    if os.path.exists(db_path):
        os.remove(db_path)
    
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS products (
        sku TEXT PRIMARY KEY,
        name TEXT,
        category TEXT,
        subcategory TEXT,
        brand TEXT,
        product_type TEXT,
        specification TEXT,
        unit TEXT,
        moq INTEGER,
        market_price REAL,
        discount REAL,
        chittety_price REAL,
        availability TEXT,
        price_basis TEXT,
        short_description TEXT,
        code_prefix TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS services (
        code TEXT PRIMARY KEY,
        category TEXT,
        name TEXT,
        starting_price REAL,
        price_type TEXT,
        notes TEXT,
        cta TEXT,
        source TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS quote_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        phone TEXT,
        email TEXT,
        company TEXT,
        project_type TEXT,
        category TEXT,
        sku TEXT,
        quantity INTEGER,
        delivery_location TEXT,
        message TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # Insert products
    for p in products:
        c.execute('''INSERT OR REPLACE INTO products 
            (sku, name, category, subcategory, brand, product_type, specification, 
             unit, moq, market_price, discount, chittety_price, availability, 
             price_basis, short_description, code_prefix)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (p['sku'], p['name'], p['category'], p['subcategory'], p['brand'],
             p['productType'], p['specification'], p['unit'], p['moq'],
             p['marketPrice'], p['discount'], p['chittetyPrice'],
             p['availability'], p['priceBasis'], p['shortDescription'],
             p['codePrefix']))
    
    # Insert services
    for s in services:
        c.execute('''INSERT OR REPLACE INTO services 
            (code, category, name, starting_price, price_type, notes, cta, source)
            VALUES (?,?,?,?,?,?,?,?)''',
            (s['code'], s['category'], s['name'], s['startingPrice'],
             s['priceType'], s['notes'], s['cta'], s['source']))
    
    # Create indexes
    c.execute('CREATE INDEX idx_products_category ON products(category)')
    c.execute('CREATE INDEX idx_products_subcategory ON products(subcategory)')
    c.execute('CREATE INDEX idx_products_brand ON products(brand)')
    c.execute('CREATE INDEX idx_products_code_prefix ON products(code_prefix)')
    c.execute('CREATE INDEX idx_services_category ON services(category)')
    
    conn.commit()
    conn.close()
    
    print(f"\nDatabase created at {db_path}")
    print("Done!")

if __name__ == '__main__':
    import_products_and_services()
